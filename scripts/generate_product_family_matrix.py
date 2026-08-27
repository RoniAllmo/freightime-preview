#!/usr/bin/env python3
"""Development-time conversion: FreighTime_Simple_Import_Requirements_Matrix.xlsx
-> js/import-readiness/product-family-matrix.js

This script runs ONLY at development time, on request, when the product
owner updates the workbook. It is never invoked from the browser and is
not part of the production runtime -- the generated output is a plain
frozen JS data module with no Excel-parsing code anywhere near it.

Usage:
    python3 scripts/generate_product_family_matrix.py

Requires: openpyxl (pip install openpyxl). Not a project dependency --
this repository ships no package.json/build step; this script is a
standalone maintenance tool, matching the "development import or
generation command" requested for this feature.

Interpretation rules (must match docs/product-family-matrix-engine.md):
  - "כן"    -> true  (a positive regulatory signal for that category)
  - "לבדוק" -> false (NOT "unknown" -- converted to a definite false;
                       "לבדוק" never survives into runtime as a pending state)
  - blank note fields stay blank; no instructional text is invented here
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("This script requires openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = REPO_ROOT / "data" / "FreighTime_Simple_Import_Requirements_Matrix.xlsx"
OUTPUT_PATH = REPO_ROOT / "js" / "import-readiness" / "product-family-matrix.js"
SHEET_NAME = "דרישות יבוא"

EXPECTED_HEADERS = [
    "קטגוריה",
    "מוצר / משפחת טובין",
    "תקינה",
    "משרד הבריאות",
    "משרד התחבורה / מעבדת רכב",
    "משרד התקשורת",
    "משרד החקלאות",
    "אישור / רישיון אחר",
    "הערה ליבוא אישי",
    "הערה ליבוא מסחרי",
    "מענה במערכת",
    "הערות קצרות",
]

REGULATORY_COLUMN_TO_SIGNAL_KEY = {
    "תקינה": "standards",
    "משרד הבריאות": "healthUmbrella",
    "משרד התחבורה / מעבדת רכב": "transportOrVehicleLaboratory",
    "משרד התקשורת": "communications",
    "משרד החקלאות": "agriculture",
    "אישור / רישיון אחר": "otherPermit",
}

COVERAGE_MAP = {
    "קיים": "existing",
    "חלקי": "partial",
    "חסר": "missing",
}

CATEGORY_SLUG = {
    "מזון ומשקאות": "food-and-beverages",
    "מגע עם מזון": "food-contact",
    "חשמל ואלקטרוניקה": "electrical-and-electronics",
    "רכב ותחבורה": "vehicles-and-transport",
    "ילדים ותינוקות": "children-and-infants",
    "בריאות ותמרוקים": "health-and-cosmetics",
    "כימיקלים וחומרים": "chemicals-and-materials",
    "טקסטיל וריהוט": "textiles-and-furniture",
    "בנייה ותעשייה": "construction-and-industrial",
    "מוצרי צריכה נוספים": "additional-consumer-products",
    "אחר": "other",
}

# Curated, explicitly-reviewed alias additions -- deliberately small and
# literal. Per the task's Phase H.57 boundary ("do not invent broad
# synonyms that change the professional scope"), every alias here is a
# direct linguistic variant of the family name itself, not a new
# professional interpretation.
CURATED_ALIASES = {
    "מזון ארוז": [
        "מזון ארוז", "אוכל ארוז", "מוצר מזון ארוז", "חטיף ארוז",
        # Wave 1 (product-owner approved, 2026-08-25): canned/preserved
        # food terms -- deliberately excludes "שימורי דגים"/"שימורי בשר",
        # which collide with the existing "דגים"/"בשר" aliases on the
        # food-of-animal-origin family below and are deferred pending a
        # separate product-owner precedence decision.
        "שימורים", "שימורי ירקות", "קופסת שימורים", "מזון משומר", "canned food", "canned goods",
    ],
    "משקאות": ["משקאות", "משקה", "משקה ארוז"],
    # Wave 2 completion (product-owner approved, 2026-08-26): the bare
    # "ויטמינים" alias was REMOVED here -- it collided with the two new
    # rows below ("ויטמינים לבעלי חיים", "ויטמינים לייצור תרופות"), which
    # both contain "ויטמינים" as a substring, making every animal-use or
    # pharmaceutical-manufacturing-use description ALSO match this human
    # row and become falsely ambiguous. Replaced with human-use-specific
    # compound phrases so real human-use text still resolves cleanly, and
    # a genuinely bare "ויטמינים"/"vitamin product"/"supplement"
    # description (no distinguishing intended-use wording) now correctly
    # stays unresolved rather than defaulting to the human-health
    # authority, per the approved rule.
    "תוספי תזונה": [
        "תוספי תזונה", "תוסף תזונה", "תוסף מזון", "תוסף תזונה לאדם",
        "ויטמינים למאכל אדם", "vitamins for human consumption",
    ],
    "מזון מן החי": ["מזון מן החי", "בשר", "עוף", "דגים", "מוצרי חלב", "גבינה", "דבש", "ביצים", "ביצים טריות", "ביצי מאכל", "מוצרי ביצים"],
    "תוצרת חקלאית, זרעים וצמחים": ["תוצרת חקלאית", "זרעים", "שתילים", "צמחים", "פירות וירקות"],
    "כלי פלסטיק במגע עם מזון": [
        "כלי פלסטיק במגע עם מזון", "קופסת פלסטיק לאוכל", "כלי אחסון מזון מפלסטיק",
        # Wave 1 (product-owner approved, 2026-08-25): colloquial phrasing variants.
        "כלי פלסטיק למזון", "קופסת אוכל",
        # Wave 3 (product-owner approved, 2026-08-27): a plastic bottle
        # intended for food/beverage contact is this same food-contact-01
        # concept -- narrow compound phrases only, never bare "בקבוק".
        "בקבוק פלסטיק למזון", "בקבוק פלסטיק למשקה",
    ],
    "מוצר עם ציפוי פולימרי במגע עם מזון": [
        "ציפוי פולימרי במגע עם מזון", "מחבת עם ציפוי לא נדבק", "כלי בישול מצופה",
        # Wave 3 (product-owner approved, 2026-08-27): polymer-coated
        # CARDBOARD packaging intended for food contact is the same
        # existing concept as any other polymer-coated food-contact
        # article -- narrow compound phrases only, never bare "קרטון".
        "קרטון מצופה פולימר למגע עם מזון", "אריזת קרטון בציפוי פולימרי למזון",
        "polymer-coated cardboard for food contact",
    ],
    "כלי זכוכית במגע עם מזון או שתייה": [
        "כלי זכוכית במגע עם מזון", "כוס זכוכית", "כלי זכוכית לשתייה", "צנצנת זכוכית למזון",
        # Wave 3 (product-owner approved, 2026-08-27): a glass bottle
        # intended for food/beverage contact is this same existing
        # concept -- narrow compound phrases only, never bare "בקבוק"
        # (a decorative or non-food-contact glass bottle must not match).
        "בקבוק זכוכית למשקה", "בקבוק זכוכית למזון", "בקבוק הבא במגע עם משקה",
        "food-contact bottle", "beverage bottle", "glass food jar", "glass drinking vessel",
    ],
    "כלי קרמיקה במגע עם מזון": [
        "כלי קרמיקה במגע עם מזון", "צלחת קרמיקה", "כוס קרמיקה",
        # Wave 3 (product-owner approved, 2026-08-27): additional
        # ceramic food-contact tableware phrasing.
        "קערת קרמיקה", "ספל קרמיקה", "ceramic plate", "ceramic bowl", "ceramic mug",
    ],
    "כלי מתכת במגע עם מזון": ["כלי מתכת במגע עם מזון", "סיר מתכת", "מחבת מתכת"],
    "מכשיר חשמלי עם תקע או ספק כוח": ["מכשיר חשמלי", "מכשיר עם תקע", "מכשיר עם ספק כוח"],
    "מוצר אלחוטי, Wi-Fi או Bluetooth": [
        "מוצר אלחוטי", "Wi-Fi", "Bluetooth", "בלוטות", "וויי פיי", "מוצר עם קליטה אלחוטית",
        "ווקי טוקי", "מכשיר קשר", "מכשירי קשר", "רדיו דו כיווני", 'מקמ"ש', "walkie talkie", "two-way radio",
    ],
    "ציוד סלולרי ותקשורת": ["ציוד סלולרי", "ציוד תקשורת", "מכשיר סלולרי"],
    "פנסים וגופי תאורה לרכב": [
        "פנס לרכב", "פנסים לרכב", "גוף תאורה לרכב", "פנס ראש לרכב",
        "פנס קדמי לרכב", "פנס אחורי לרכב", "פנס ראשי לרכב", "פנס איתות לרכב",
    ],
    # Wave 2 completion (product-owner approved, 2026-08-26): this row
    # previously named "תמרוקים ובשמים" and bundled perfume with
    # cosmetics under one shared positive signal. Split into two rows
    # (this one, cosmetics-only; "בשמים" below, perfume-only) because the
    # approved rule requires two DIFFERENT primary directions. Renamed to
    # "תמרוקים" -- "בושם"/"perfume"/etc. removed from this row's aliases
    # (moved to the new perfume row below). The legacy compound name
    # "תמרוקים ובשמים" is deliberately NOT re-added as an alias here: it
    # contains "בשמים", the new perfume row's own base name, so it would
    # make a handful of pre-existing tests' literal input text genuinely
    # (and correctly) ambiguous between both rows. Those tests were
    # updated to use an unambiguous term ("קוסמטיקה"/"תמרוקים") instead.
    "תמרוקים": [
        "תמרוקים", "קוסמטיקה", "מוצר קוסמטי",
        "לק ג'ל", "ג'ל לק", "לק לציפורניים", "תמרוק", "קרם פנים",
        "דיאודורנט", "deodorant", "קרם לטיפול בעור", "תכשיר לשיער",
        "תכשיר איפור", "makeup product",
        # Wave 1 (product-owner approved, 2026-08-25): moisturizer variant.
        # Deliberately excludes bare "קרם" -- collides with food terms
        # ("קרם גבינה", "קרם עוגה").
        "קרם לחות",
    ],
    # New row (Wave 2 completion): perfume, split off from the cosmetics
    # row above. No positive signal (see the workbook: healthUmbrella is
    # "לבדוק" for this row) -- the approved rule states perfume generally
    # does not require import approval, unlike cosmetics. Deliberately
    # uses ONLY the singular "בושם", never the plural "בשמים" -- see the
    # note on the cosmetics row above for why.
    "בשמים": ["בושם", "perfume", "eau de parfum", "eau de toilette"],
    "ציוד רפואי": ["ציוד רפואי", "מכשור רפואי"],
    "מוצר בעל טענה רפואית": ["טענה רפואית", "מוצר עם טענה רפואית"],
    "ביגוד וטקסטיל": [
        "ביגוד", "בגד", "טקסטיל", "בגדים",
        # Wave 1 (product-owner approved, 2026-08-25): garment-type
        # terms. Deliberately excludes "מגבת"/"towel" -- ambiguous with
        # a cleaning cloth ("מגבת ניקוי"), deferred pending review.
        # "dress" was approved but is OMITTED here: code review found it
        # is a plain substring of common unrelated English words --
        # "address", "dresser", "dressing", "redress", "dressage",
        # "undressed" -- so any commercial-import text mentioning a
        # shipping address, a piece of furniture, or a food dressing
        # would be misidentified as this family. The Hebrew alias
        # "שמלה" already covers the dress concept safely; omitting the
        # unsafe English alias rather than modifying the substring-match
        # architecture, per this PR's explicit safety boundary.
        "חולצה", "מכנס", "מכנסיים", "שמלה", "ג'קט", "מעיל", "גרביים", "גרב",
        "shirt", "t-shirt",
    ],
    # Animal-feed completion (2026-08-27, product-owner approved):
    # "מזון לחיות מחמד" (pet food) removed from this row -- animal feed
    # is a separate concept from non-food pet products/accessories and
    # now routes to the dedicated "מזון לבעלי חיים" row below instead.
    # Wave 3 (product-owner approved, 2026-08-27): non-food pet
    # equipment/accessories added -- narrow, exact compound phrases
    # only. This row's own regulatorySignals were also corrected this
    # pass (see the workbook) -- non-food pet products generally
    # require no positive regulatory approval, not the Ministry of
    # Health direction the row previously carried by mistake.
    "מוצרים לבעלי חיים": [
        "מוצר לבעלי חיים", "מוצרים לחיות מחמד",
        "רצועה לכלב", "קולר לחתול", "מיטה לחיית מחמד", "קערה לחיית מחמד",
        "צעצוע לחיית מחמד", "מברשת טיפוח לחיית מחמד", "אביזר לאקווריום",
        "pet leash", "pet collar", "pet bed", "pet bowl", "pet toy",
        "grooming brush for pets", "aquarium accessory",
    ],
    # Wave 1 (product-owner approved, 2026-08-25): this family previously
    # had no curated aliases beyond its own name.
    "חומרי ניקוי וחיטוי": ["חומר ניקוי", "נוזל ניקוי", "אבקת כביסה", "cleaning product"],

    # ---------------------------------------------------------------
    # Wave 2 completion (product-owner approved, 2026-08-26)
    # ---------------------------------------------------------------

    # Renamed from "הנעלה" (ordinary+safety footwear, one shared "no
    # signal" row) -- split into this row (ordinary, unchanged no
    # signal) and "הנעלת בטיחות" below (new row, positive standards
    # signal), per the approved rule that safety footwear needs a
    # different direction than ordinary footwear.
    # Deliberately excludes bare "נעל" (singular "shoe") -- it is a
    # substring of the safety-footwear row's own aliases ("נעלי בטיחות",
    # "הנעלת בטיחות"), which would make safety-footwear text also match
    # this row and become falsely ambiguous. "נעליים" (plural) does not
    # have this problem. See also FAMILY_NEGATIVE_TERMS in
    # product-family-identification.js for the English "shoes"/"boots"
    # equivalent exclusion.
    "הנעלה רגילה": [
        "הנעלה", "נעליים", "מגפיים", "סנדלים", "נעלי ספורט",
        "shoes", "boots", "sandals", "sports shoes",
    ],
    # New row: safety footwear, split off from "הנעלה" above.
    "הנעלת בטיחות": [
        "נעלי בטיחות", "נעלי עבודה עם מיגון", "safety shoes", "safety boots", "protective footwear",
    ],
    # Renamed from "ציוד ספורט וציוד מגן" (sports+protective equipment,
    # one shared "no signal" row) -- split into this row (sports,
    # unchanged no signal) and "ציוד מגן אישי" below (new row, positive
    # standards signal), per the approved rule that protective equipment
    # needs a different direction than ordinary sports equipment.
    # Code-review finding, fixed: the legacy compound name "ציוד ספורט
    # וציוד מגן" is deliberately NOT re-added as an alias here (same
    # reasoning as "תמרוקים ובשמים" on the cosmetics/perfume split
    # above) -- it explicitly names protective equipment too, so
    # resolving it to this no-signal row would falsely tell a genuinely
    # protective-equipment description that no approval is needed.
    "ציוד ספורט": [
        "משקולות", "ציוד ספורט", "מכשיר אימון לא חשמלי", "non-electric training equipment",
        "weights", "ordinary sports equipment",
    ],
    # New row: personal protective equipment, split off from "ציוד ספורט
    # וציוד מגן" above.
    "ציוד מגן אישי": [
        "קסדת מגן", "משקפי מגן", "כפפות מגן", "רתמת בטיחות", "ציוד הגנה נשימתית",
        "protective helmet", "protective eyewear", "protective gloves",
        "respiratory protection equipment", "safety harness",
        "sport protective equipment", "work protective equipment",
    ],
    # Renamed from "אופניים וקורקינטים" (ordinary+auxiliary-motor, one
    # shared positive-standards row) -- split into this row (ordinary,
    # unchanged standards signal) and "אופניים או קורקינט עם מנוע עזר"
    # below (new row, transportOrVehicleLaboratory signal instead), per
    # the approved rule that a motorized bicycle/scooter needs the
    # certified-vehicle-laboratory route instead of a plain standards
    # check. See FAMILY_NEGATIVE_TERMS in product-family-identification.js
    # for how accessory phrasing ("מנשא אופניים לרכב") is protected from
    # being forced into either of these two families.
    "אופניים וקורקינטים רגילים": [
        "אופניים וקורקינטים", "אופניים", "אופני הרים", "אופני ילדים", "קורקינט רגיל",
        "bicycle", "mountain bicycle", "non-motorized scooter",
    ],
    # New row: bicycle/scooter with an auxiliary motor, split off from
    # "אופניים וקורקינטים" above.
    "אופניים או קורקינט עם מנוע עזר": [
        "אופניים חשמליים", "אופניים עם מנוע עזר", "קורקינט חשמלי", "קורקינט עם מנוע עזר",
        "electric bicycle", "bicycle with auxiliary motor", "electric scooter", "scooter with auxiliary motor",
    ],
    # New row: vitamins for animal consumption -- the pre-existing
    # "תוספי תזונה" row (human-use food supplements/vitamins) is left
    # completely unchanged; this is a genuinely separate product concept
    # requiring the Veterinary Services direction (agriculture signal)
    # instead of the Ministry of Health (healthUmbrella signal).
    "ויטמינים לבעלי חיים": [
        "ויטמינים לבעלי חיים", "תוסף ויטמינים לבעלי חיים", "animal vitamins", "veterinary feed vitamins",
    ],
    # New row: vitamins intended as a pharmaceutical-manufacturing raw
    # material -- a genuinely separate product concept from an
    # end-consumer food supplement, even though both ultimately route
    # through the Ministry of Health (healthUmbrella); the family-guidance
    # overlay (product-family-guidance.js) supplies the more specific
    # "אגף הרוקחות" (Pharmaceutical Division) wording for this row only.
    "ויטמינים לייצור תרופות": [
        "ויטמינים לייצור תרופות", "חומר גלם ויטמיני לייצור תרופות", "vitamins for pharmaceutical manufacturing",
    ],

    # ---------------------------------------------------------------
    # Final completion pass (product-owner approved, 2026-08-26):
    # batteries/accumulators and the furniture/mattress split.
    # ---------------------------------------------------------------

    # Batteries and accumulators as the imported product (existing
    # positive `standards` signal, row/id unchanged). Deliberately
    # excludes bare "מצבר" would collide with the new vehicle-dedicated
    # row's own "מצבר לרכב" phrasing (accumulator generically still
    # matches here; the vehicle-specific compound is excluded via
    # FAMILY_NEGATIVE_TERMS in product-family-identification.js, not by
    # omitting "מצבר" here -- omitting it would also break the required
    # bare "accumulator"/"מצבר" standalone scenario).
    "סוללות ותאים": [
        "סוללה", "סוללת ליתיום", "מארז סוללות", "מצבר",
        "battery", "lithium battery", "battery pack", "accumulator",
    ],
    # Renamed from "ריהוט ומזרנים" (furniture+mattresses, one shared
    # positive-standards row) -- split into this row (mattresses,
    # unchanged standards signal) and "ריהוט" below (new row, no
    # positive signal), per the approved rule that ordinary furniture
    # without electrical wiring generally has no positive direction.
    # Deliberately does NOT re-add the legacy compound "ריהוט ומזרנים"
    # as an alias -- same reasoning as the cosmetics/perfume and
    # sports/protective-equipment splits above (a code-review finding on
    # the sports split): it explicitly also names furniture, so
    # resolving it to this row would falsely imply a positive direction
    # for text that may only mean ordinary furniture.
    "מזרנים": ["מזרן", "mattress", "mattresses"],
    # New row: ordinary furniture, split off from "ריהוט ומזרנים" above.
    # No positive signal -- electrically wired furniture instead reuses
    # the existing, family-independent mains-connected-electrical-product
    # detailed rule (see product-family-wave2-completion.test.js for the
    # regression test), not a matrix signal on this row.
    "ריהוט": [
        "שולחן", "כיסא", "ארון", "שידה", "ספה",
        "table", "chair", "cabinet", "ordinary furniture",
    ],
    # New row: a vehicle-dedicated accumulator, split off conceptually
    # from the standalone-battery row above -- a genuinely different
    # required direction (certified vehicle laboratory, not a plain
    # standards check). Only `transportOrVehicleLaboratory` is positive
    # here (not `standards`) so exactly one CTA/professional shows,
    # per the approved rule's "do not show two equal CTAs" requirement
    # -- selectPrimaryAndSupportingProfessional (product-family-result.js)
    # would otherwise prefer `standards` over `transportOrVehicleLaboratory`
    # by SIGNAL_ORDER precedence if both were set.
    "מצבר ייעודי לרכב": [
        "מצבר לרכב", "vehicle battery", "car battery", "vehicle accumulator",
    ],
    # New row (grouped-battery-selection completion, 2026-08-26):
    # equipment where a battery is merely a component, not the imported
    # product itself -- genuinely different from the standalone-battery
    # row above (no positive signal here; a Standards Institution
    # category is never fabricated just because equipment happens to
    # contain a battery). See the batteries_or_battery_containing
    # checkbox change in product-family-selection-mapping.js and
    # FAMILY_NEGATIVE_TERMS in product-family-identification.js, which
    # excludes this same wording from the standalone-battery row.
    "ציוד הכולל סוללה": [
        "מוצר הכולל סוללה פנימית", "ציוד נייד עם סוללה",
        "equipment containing an internal battery", "battery-powered device", "rechargeable device", "rechargeable equipment",
    ],
    # New row (live-animals completion, 2026-08-26): a live animal
    # itself, distinct from the pre-existing "מזון מן החי" (products OF
    # animal origin) and "מוצרים לבעלי חיים"/"ויטמינים לבעלי חיים"
    # (products/vitamins FOR animals) rows. Deliberately narrow -- only
    # "בעל חיים"/"בעלי חיים"/"live animal"/"live animals" -- per the
    # product owner's explicit rule not to add broad animal-species
    # aliases. See FAMILY_NEGATIVE_TERMS in
    # product-family-identification.js, which excludes the "ל...בעלי
    # חיים" ("for animals") prepositional phrasing so this row's own
    # plural alias can never falsely match the pre-existing
    # products-for-animals/vitamins-for-animals rows' own aliases (both
    # contain "בעלי חיים" as a substring).
    "בעלי חיים": ["בעל חיים", "live animal", "live animals"],
    # New row (animal-feed completion, 2026-08-27): feed intended as
    # food for animals -- distinct from live animals themselves
    # ("בעלי חיים" above), products OF animal origin ("מזון מן החי"),
    # and non-food pet products/accessories ("מוצרים לבעלי חיים",
    # unchanged). Deliberately narrow, exact compound phrases only --
    # per the product owner's explicit rule against bare "מזון"/"food"/
    # "feed"/"pet"/species-name aliases, which would falsely match
    # unrelated food and non-food products. See
    # FAMILY_NEGATIVE_TERMS in product-family-identification.js for the
    # non-food pet-product protection.
    "מזון לבעלי חיים": [
        "מזון לכלבים", "מזון לחתולים", "מזון לדגים", "מזון לציפורים", "מזון לחיות משק",
        "מזון לחיות מחמד",
        "animal feed", "dog food", "cat food", "fish food", "bird food", "livestock feed", "pet food",
    ],

    # ---------------------------------------------------------------
    # Wave 3 (product-owner approved, 2026-08-27)
    # ---------------------------------------------------------------

    # Toys (existing row, pre-existing no curated aliases beyond its
    # own name "צעצועים"): a children's book that IS a toy/play product
    # -- narrow, exact compound phrases only, per the explicit rule
    # that an ordinary book must never become a toy merely because it
    # is a children's book. Genuinely ambiguous "ספר ילדים" (children's
    # book) text alone stays unresolved -- not added here.
    "צעצועים": [
        "ספר צעצוע", "ספר פעילות עם חלקי משחק", "ספר ילדים בעל ערך משחקי",
        "interactive play book for children", "toy book",
    ],
    # Vehicle glass/windshields (existing row): additional vehicle
    # safety-glass phrasing -- vehicle installation route takes
    # precedence over the new building-safety-glass row below (this
    # row is unaffected by that addition; the two are disambiguated by
    # their own distinct "לרכב" vs "לבניין" compound phrasing).
    "זכוכית ושמשות לרכב": [
        "זכוכית בטיחות לרכב", "זכוכית בטחון להתקנה ברכב", "vehicle safety glass", "automotive safety glass",
    ],
    # Drones (new row): Ministry of Communications route, reached via
    # the existing wireless_or_transmitting_equipment checkbox
    # candidate set. Deliberately excludes bare "רחפן" fragments that
    # only name an accessory/part, not the complete product -- see
    # FAMILY_NEGATIVE_TERMS for the accessory exclusion.
    "רחפן": ["drone", "camera drone", "commercial drone"],
    # Hand tools, ordinary (new row, no positive signal): a hand tool
    # that does not connect to mains power. The mains-connected case
    # reuses the existing, family-independent mains-connected-
    # electrical-product detailed rule -- no data needed here for that
    # case, it fires independently of this row.
    "כלי עבודה ידניים": [
        "פטיש", "מברג", "פלייר", "מפתח ברגים", "מסור ידני", "סט כלי עבודה ידניים",
        "hammer", "screwdriver", "pliers", "wrench", "hand saw", "hand tool set",
    ],
    # Ordinary cardboard packaging (new row, no positive signal):
    # deliberately excludes bare "קרטון"/"cardboard"/"box" -- see the
    # polymer-coated food-contact cardboard aliases added to
    # food-contact-02 above, which must take precedence for that
    # narrower concept.
    "קרטון לאריזה": [
        "קופסת קרטון", "ארגז קרטון", "קרטון גלי",
        "cardboard box", "corrugated carton", "shipping carton",
    ],
    # Wooden boxes for packaging (new row, Ministry of Agriculture
    # route -- reuses the same existing authority already used for
    # live animals/animal feed/products of animal origin, no new
    # professional category).
    "קופסת עץ לאריזה": ["ארגז עץ", "wooden box", "wooden crate"],
    # Ordinary paper and printed products (new row, no positive
    # signal): deliberately excludes bare "ספר"/"נייר"/"book"/"paper"
    # -- the toys row above already carries the narrow play-value-book
    # exception, disambiguated by its own distinct compound phrasing.
    "נייר ומוצרי דפוס": [
        "נייר להדפסה", "מחברת", "ספר רגיל", "מדבקות", "תוויות", "שקית נייר", "נייר תרמי", "מוצר דפוס",
        "printing paper", "notebook", "ordinary book", "printed product", "paper label",
    ],
    # Carpets and rugs (new row, Standards Institution). Deliberately
    # excludes "carpet cleaner"/"carpet cleaning" phrasing -- see
    # FAMILY_NEGATIVE_TERMS.
    "שטיחים": ["שטיח", "rug", "carpet"],
    # Ordinary blanket, no electrical wiring (new row, no positive
    # signal). The electrically wired case reuses the existing
    # mains-connected-electrical-product detailed rule, unaffected by
    # this row.
    "שמיכה רגילה": ["שמיכה", "שמיכת בד", "blanket", "textile blanket"],
    # Pacifier holder (new row, Standards Institution, reached via the
    # existing childrens_products_and_toys checkbox candidate set).
    "מחזיק מוצץ": ["pacifier holder", "pacifier clip"],
    # Infant carrier (new row, Standards Institution, reached via the
    # existing childrens_products_and_toys checkbox candidate set).
    # Deliberately excludes bare "מנשא"/"carrier" -- see
    # FAMILY_NEGATIVE_TERMS protecting against a vehicle-carrier
    # collision.
    "מנשא לתינוק": ["baby carrier", "infant carrier"],
    # Ordinary non-apparel textile products (new row, no positive
    # signal) -- bedding, curtains, towels, upholstery fabric, fabric
    # bags. Deliberately excludes bare "בד"/"fabric"/"textile". Named
    # "מוצרי טקסטיל ביתיים" (household textile products) rather than
    # "...שאינם ביגוד" ("...that are not apparel") -- the negated form
    # would have contained "ביגוד" as a plain substring of its own
    # name, colliding with the existing apparel row's own "ביגוד"
    # alias (code-review-caught during this pass's own collision scan).
    "מוצרי טקסטיל ביתיים": [
        "מצעים", "וילונות", "מגבות", "בד לריפוד", "תיק בד",
        "bedding", "curtains", "towels", "upholstery fabric", "fabric bag",
    ],
    # Building safety glass (new row, Standards Institution) -- distinct
    # from the existing building-materials no-positive row (a genuinely
    # different required direction) and from vehicle safety glass
    # (existing "זכוכית ושמשות לרכב" row, vehicle-laboratory route).
    "זכוכית בטיחות לבניין": [
        "זכוכית בטחון לשימוש בבניינים", "building safety glass", "architectural safety glass",
    ],
}


def normalize(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def convert_regulatory_value(raw_value, *, column, row_number, errors):
    text = normalize(raw_value)
    if text == "כן":
        return True
    if text == "לבדוק":
        return False
    errors.append(
        f"Row {row_number}: unsupported value {raw_value!r} in column {column!r} "
        "(expected 'כן' or 'לבדוק')"
    )
    return False


def slugify_family_name(category_slug, index_in_category):
    return f"{category_slug}-{index_in_category:02d}"


def build_aliases(family_name):
    aliases = [family_name]
    for extra in CURATED_ALIASES.get(family_name, []):
        if extra not in aliases:
            aliases.append(extra)
    return aliases


def load_rows():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found at {WORKBOOK_PATH}")
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(
            f"Expected sheet {SHEET_NAME!r}, found sheets: {workbook.sheetnames}"
        )
    sheet = workbook[SHEET_NAME]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize(h) for h in header_row]
    if headers != EXPECTED_HEADERS:
        raise SystemExit(
            "Workbook header row does not match the expected columns.\n"
            f"Expected: {EXPECTED_HEADERS}\nFound:    {headers}"
        )
    rows = list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True))
    return headers, rows


def build_registry():
    headers, rows = load_rows()
    idx = {h: i for i, h in enumerate(headers)}
    errors = []
    seen_names = set()
    duplicate_names = set()
    category_counters = {}
    records = []

    stats = {
        "rowCount": len(rows),
        "categoryCount": 0,
        "yesCountsByCategory": {v: 0 for v in REGULATORY_COLUMN_TO_SIGNAL_KEY.values()},
        "coverageCounts": {"existing": 0, "partial": 0, "missing": 0},
        "blankPersonalNotes": 0,
        "blankCommercialNotes": 0,
        "duplicateFamilyNames": [],
        "malformedRows": [],
        "unsupportedValues": len(errors),
        "placeholderRowsExcluded": 0,
    }

    categories_seen = set()

    for offset, row in enumerate(rows):
        row_number = offset + 2  # workbook row number, 1-indexed with header
        category = normalize(row[idx["קטגוריה"]])
        family_name = normalize(row[idx["מוצר / משפחת טובין"]])

        if not category or not family_name:
            stats["malformedRows"].append(row_number)
            continue

        categories_seen.add(category)

        if family_name in seen_names:
            duplicate_names.add(family_name)
        seen_names.add(family_name)

        signals = {}
        for column, signal_key in REGULATORY_COLUMN_TO_SIGNAL_KEY.items():
            value = convert_regulatory_value(
                row[idx[column]], column=column, row_number=row_number, errors=errors
            )
            signals[signal_key] = value
            if value:
                stats["yesCountsByCategory"][signal_key] += 1

        personal_note = normalize(row[idx["הערה ליבוא אישי"]])
        commercial_note = normalize(row[idx["הערה ליבוא מסחרי"]])
        if personal_note is None:
            stats["blankPersonalNotes"] += 1
        if commercial_note is None:
            stats["blankCommercialNotes"] += 1

        coverage_raw = normalize(row[idx["מענה במערכת"]])
        coverage = COVERAGE_MAP.get(coverage_raw)
        if coverage is None:
            errors.append(
                f"Row {row_number}: unsupported מענה במערכת value {coverage_raw!r}"
            )
            coverage = "missing"
        stats["coverageCounts"][coverage] += 1

        short_notes = normalize(row[idx["הערות קצרות"]])

        # Row 52 in the source workbook is a manual-completion placeholder
        # ("משפחה נוספת להשלמה ידנית") with no real product-owner content --
        # every regulatory column is "לבדוק" and both notes are blank. It is
        # excluded from the active registry (activeStatus:false) rather than
        # invented into a real family.
        is_placeholder = family_name == "משפחה נוספת להשלמה ידנית" and category == "אחר"
        if is_placeholder:
            stats["placeholderRowsExcluded"] += 1

        category_slug = CATEGORY_SLUG.get(category)
        if category_slug is None:
            errors.append(f"Row {row_number}: unmapped category {category!r}")
            category_slug = "uncategorized"
        category_counters[category_slug] = category_counters.get(category_slug, 0) + 1
        family_id = slugify_family_name(category_slug, category_counters[category_slug])

        records.append(
            {
                "id": family_id,
                "category": category,
                "publicFamilyName": family_name,
                "aliases": build_aliases(family_name),
                "regulatorySignals": signals,
                "personalImportNote": personal_note,
                "commercialImportNote": commercial_note,
                "currentSystemCoverage": coverage,
                "shortNotes": short_notes,
                "optionalSubdomain": None,
                "activeStatus": not is_placeholder,
                "version": 1,
                "productOwnerReviewedDate": "2026-08-18",
                "sourceRow": row_number,
            }
        )

    stats["categoryCount"] = len(categories_seen)
    stats["duplicateFamilyNames"] = sorted(duplicate_names)

    if errors:
        raise SystemExit("Workbook validation failed:\n" + "\n".join(errors))

    if duplicate_names:
        raise SystemExit(f"Duplicate family names rejected: {sorted(duplicate_names)}")

    return records, stats


JS_HEADER = """/**
 * GENERATED FILE -- do not hand-edit.
 *
 * Produced by scripts/generate_product_family_matrix.py from the
 * product-owner-authored workbook at
 * data/FreighTime_Simple_Import_Requirements_Matrix.xlsx (sheet
 * \"דרישות יבוא\").
 * See docs/product-family-matrix-engine.md for the full interpretation
 * rules and the update workflow.
 *
 * Interpretation already applied during generation (not at runtime):
 *   \"כן\"    -> true  (positive regulatory signal)
 *   \"לבדוק\" -> false (definite negative signal, not \"unknown\")
 *
 * Regenerate with: python3 scripts/generate_product_family_matrix.py
 */

export const PRODUCT_FAMILY_MATRIX = Object.freeze(
"""

JS_FOOTER = """);

export function findFamilyById(id) {
  return PRODUCT_FAMILY_MATRIX.find((family) => family.id === id) || null;
}

export function activeFamilies() {
  return PRODUCT_FAMILY_MATRIX.filter((family) => family.activeStatus);
}

export function familiesByCategory(category) {
  return activeFamilies().filter((family) => family.category === category);
}
"""


def render_js(records):
    # json.dumps with ensure_ascii=False keeps Hebrew readable in the
    # generated source, and .map(...) below freezes every nested object
    # so the registry is deeply immutable, not just the top-level array.
    body = json.dumps(records, ensure_ascii=False, indent=2)
    js = (
        JS_HEADER
        + "  "
        + body.replace("\n", "\n  ")
        + ".map((family) => Object.freeze({ ...family, regulatorySignals: Object.freeze({ ...family.regulatorySignals }), aliases: Object.freeze([...family.aliases]) }))"
        + JS_FOOTER
    )
    return js


def main():
    records, stats = build_registry()
    js = render_js(records)
    OUTPUT_PATH.write_text(js, encoding="utf-8")
    print(f"Wrote {len(records)} family records to {OUTPUT_PATH}")
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
